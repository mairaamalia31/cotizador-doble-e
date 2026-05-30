from flask import Flask, render_template, request, jsonify, send_file
import json, os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import io
import psycopg2
from psycopg2.extras import RealDictCursor
import urllib.parse

app = Flask(__name__)

def get_db():
    url = os.environ.get("DATABASE_URL")
    if not url:
        raise Exception("DATABASE_URL no configurada")
    result = urllib.parse.urlparse(url)
    conn = psycopg2.connect(
        database=result.path[1:],
        user=result.username,
        password=result.password,
        host=result.hostname,
        port=result.port,
        cursor_factory=RealDictCursor,
        sslmode='require'
    )
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS cotizaciones (
        id SERIAL PRIMARY KEY,
        numero INTEGER UNIQUE,
        fecha TEXT,
        cliente TEXT,
        ruc TEXT,
        direccion TEXT,
        razon_social TEXT,
        contacto TEXT,
        telefonos TEXT,
        items TEXT,
        subtotal REAL,
        igv REAL,
        total REAL,
        tiempo_entrega TEXT,
        condicion_pago TEXT,
        created_at TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS config (
        key TEXT PRIMARY KEY,
        value TEXT
    )''')
    c.execute("INSERT INTO config (key,value) VALUES ('ultimo_numero','3126') ON CONFLICT (key) DO NOTHING")
    conn.commit()
    conn.close()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/nuevo-numero')
def nuevo_numero():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT value FROM config WHERE key='ultimo_numero'")
    val = c.fetchone()
    conn.close()
    return jsonify({'numero': int(val['value']) + 1})

@app.route('/api/cotizaciones')
def listar_cotizaciones():
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT id, numero, fecha, cliente, razon_social, contacto, total,
               tiempo_entrega, condicion_pago,
               substr(items, 1, 300) as items_preview
        FROM cotizaciones ORDER BY numero DESC
    """)
    rows = c.fetchall()
    conn.close()
    result = []
    for r in rows:
        try:
            items = json.loads(r['items_preview']) if r['items_preview'] else []
        except:
            items = []
        result.append({
            'id': r['id'], 'numero': r['numero'], 'fecha': r['fecha'],
            'cliente': r['cliente'], 'razon_social': r['razon_social'],
            'contacto': r['contacto'], 'total': r['total'],
            'tiempo_entrega': r['tiempo_entrega'], 'condicion_pago': r['condicion_pago'],
            'productos': ', '.join([i.get('detalle','') for i in items[:3]])
        })
    return jsonify(result)

@app.route('/api/cotizacion/<int:id>')
def ver_cotizacion(id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM cotizaciones WHERE id=%s", (id,))
    r = c.fetchone()
    conn.close()
    if not r:
        return jsonify({'error': 'No encontrada'}), 404
    d = dict(r)
    d['items'] = json.loads(d['items'])
    return jsonify(d)

@app.route('/api/guardar', methods=['POST'])
def guardar():
    data = request.json
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("""
            INSERT INTO cotizaciones
            (numero, fecha, cliente, ruc, direccion, razon_social, contacto, telefonos,
             items, subtotal, igv, total, tiempo_entrega, condicion_pago, created_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            data['numero'], data['fecha'], data['cliente'], data.get('ruc',''),
            data.get('direccion',''), data.get('razon_social',''), data.get('contacto',''),
            data.get('telefonos',''), json.dumps(data['items']),
            data['subtotal'], data['igv'], data['total'],
            data.get('tiempo_entrega',''), data.get('condicion_pago',''),
            datetime.now().isoformat()
        ))
        c.execute("UPDATE config SET value=%s WHERE key='ultimo_numero'", (str(data['numero']),))
        conn.commit()
        c.execute("SELECT id FROM cotizaciones WHERE numero=%s", (data['numero'],))
        new_id = c.fetchone()['id']
        conn.close()
        return jsonify({'ok': True, 'id': new_id})
    except Exception as e:
        conn.rollback()
        conn.close()
        if 'unique' in str(e).lower() or 'duplicate' in str(e).lower():
            return jsonify({'ok': False, 'error': 'Ya existe una cotización con ese número'}), 400
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/eliminar/<int:id>', methods=['DELETE'])
def eliminar(id):
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM cotizaciones WHERE id=%s", (id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/descargar/excel/<int:id>')
def descargar_excel(id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM cotizaciones WHERE id=%s", (id,))
    r = c.fetchone()
    conn.close()
    if not r:
        return "No encontrada", 404
    d = dict(r)
    d['items'] = json.loads(d['items'])
    buf = generar_excel(d)
    return send_file(buf, as_attachment=True,
                     download_name=f"Cotizacion_{d['numero']}_{d['cliente']}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/descargar/pdf/<int:id>')
def descargar_pdf(id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM cotizaciones WHERE id=%s", (id,))
    r = c.fetchone()
    conn.close()
    if not r:
        return "No encontrada", 404
    d = dict(r)
    d['items'] = json.loads(d['items'])
    buf = generar_pdf(d)
    return send_file(buf, as_attachment=True,
                     download_name=f"Cotizacion_{d['numero']}_{d['cliente']}.pdf",
                     mimetype='application/pdf')

def generar_excel(d):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotización"
    azul_oscuro = "1B2A4A"
    azul_medio = "2E4F8C"
    gris_claro = "F5F5F5"
    amarillo = "FFC107"
    blanco = "FFFFFF"
    thin = Side(style='thin', color='CCCCCC')
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    def celda(fila, col, valor, bold=False, size=11, color_fondo=None, color_texto="000000",
              alin="left", wrap=False, merge_hasta=None):
        c = ws.cell(row=fila, column=col, value=valor)
        c.font = Font(bold=bold, size=size, color=color_texto, name="Arial")
        c.alignment = Alignment(horizontal=alin, vertical="center", wrap_text=wrap)
        c.border = borde
        if color_fondo:
            c.fill = PatternFill("solid", start_color=color_fondo)
        if merge_hasta:
            ws.merge_cells(start_row=fila, start_column=col, end_row=fila, end_column=merge_hasta)
        return c

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 38
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.row_dimensions[1].height = 10

    ws.merge_cells('B2:D5')
    c = ws['B2']
    c.value = "DOBLE E PUBLICIDAD Y DISEÑO"
    c.font = Font(bold=True, size=16, color=blanco, name="Arial")
    c.fill = PatternFill("solid", start_color=azul_oscuro)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells('E2:F2')
    celda(2, 5, "N° de Cotización", bold=True, color_fondo=azul_medio, color_texto=blanco, alin="center")
    ws.merge_cells('E3:F3')
    celda(3, 5, d['numero'], bold=True, size=14, color_fondo=amarillo, alin="center")
    ws.merge_cells('E4:F4')
    celda(4, 5, "Fecha:", bold=True, color_fondo=azul_medio, color_texto=blanco, alin="center")
    ws.merge_cells('E5:F5')
    celda(5, 5, d['fecha'], color_fondo=gris_claro, alin="center")
    ws.row_dimensions[6].height = 8

    celda(7, 2, "RUC:", bold=True, color_fondo=gris_claro, merge_hasta=3)
    celda(7, 4, "10011273659", merge_hasta=6)
    celda(8, 2, "Dirección:", bold=True, color_fondo=gris_claro, merge_hasta=3)
    celda(8, 4, "Los Diamantes 387 int. 303 La Victoria Lima", merge_hasta=6)
    celda(9, 2, "Teléfono:", bold=True, color_fondo=gris_claro, merge_hasta=3)
    celda(9, 4, "Oficina 2661221 / 988455940", merge_hasta=6)
    ws.row_dimensions[10].height = 8

    celda(11, 2, "DATOS DEL CLIENTE", bold=True, size=10, color_fondo=azul_oscuro, color_texto=blanco, alin="center", merge_hasta=6)
    celda(12, 2, "Cliente:", bold=True, color_fondo=gris_claro)
    celda(12, 3, d['cliente'], merge_hasta=6)
    celda(13, 2, "RUC:", bold=True, color_fondo=gris_claro)
    celda(13, 3, d.get('ruc',''), merge_hasta=6)
    celda(14, 2, "Dirección:", bold=True, color_fondo=gris_claro)
    celda(14, 3, d.get('direccion',''), merge_hasta=6)
    celda(15, 2, "R. Social:", bold=True, color_fondo=gris_claro)
    celda(15, 3, d.get('razon_social',''), merge_hasta=6)
    celda(16, 2, "Contacto:", bold=True, color_fondo=gris_claro)
    celda(16, 3, d.get('contacto',''), merge_hasta=6)
    celda(17, 2, "Teléfonos:", bold=True, color_fondo=gris_claro)
    celda(17, 3, d.get('telefonos',''), merge_hasta=6)
    ws.row_dimensions[18].height = 8
    celda(19, 2, "En atención a su gentil solicitud, es grato cotizar a ustedes el siguiente trabajo:", color_fondo="FFF8E1", alin="left", merge_hasta=6, wrap=True)
    ws.row_dimensions[19].height = 20

    celda(20, 2, "ITEM", bold=True, color_fondo=azul_medio, color_texto=blanco, alin="center")
    celda(20, 3, "CANT.", bold=True, color_fondo=azul_medio, color_texto=blanco, alin="center")
    celda(20, 4, "DETALLE", bold=True, color_fondo=azul_medio, color_texto=blanco, alin="center")
    celda(20, 5, "PRECIO UNIT.", bold=True, color_fondo=azul_medio, color_texto=blanco, alin="center")
    celda(20, 6, "V. TOTAL", bold=True, color_fondo=azul_medio, color_texto=blanco, alin="center")

    fila_items = 21
    for i, item in enumerate(d['items'], 1):
        bg = blanco if i % 2 == 1 else gris_claro
        celda(fila_items, 2, i, color_fondo=bg, alin="center")
        celda(fila_items, 3, item.get('cantidad', 0), color_fondo=bg, alin="center")
        celda(fila_items, 4, item.get('detalle', ''), color_fondo=bg, wrap=True)
        celda(fila_items, 5, item.get('precio', 0), color_fondo=bg, alin="right")
        celda(fila_items, 6, f"=C{fila_items}*E{fila_items}", color_fondo=bg, alin="right")
        ws.row_dimensions[fila_items].height = 18
        fila_items += 1

    fr = fila_items + 1
    ws.merge_cells(f'B{fr}:E{fr}')
    celda(fr, 2, "V. VENTA (SIN IGV)", bold=True, color_fondo=gris_claro, alin="right", merge_hasta=5)
    celda(fr, 6, d['subtotal'], bold=True, alin="right")
    ws.cell(fr, 6).number_format = '#,##0.00'
    celda(fr+1, 2, "IGV (18%)", bold=True, color_fondo=gris_claro, alin="right", merge_hasta=5)
    celda(fr+1, 6, d['igv'], bold=True, alin="right")
    ws.cell(fr+1, 6).number_format = '#,##0.00'
    celda(fr+2, 2, "PRECIO TOTAL S/.", bold=True, size=12, color_fondo=azul_oscuro, color_texto=blanco, alin="right", merge_hasta=5)
    celda(fr+2, 6, d['total'], bold=True, size=12, color_fondo=amarillo, alin="right")
    ws.cell(fr+2, 6).number_format = '#,##0.00'

    fr2 = fr + 4
    if d.get('tiempo_entrega'):
        ws.merge_cells(f'B{fr2}:C{fr2}')
        celda(fr2, 2, "Tiempo de Entrega:", bold=True, color_fondo=gris_claro)
        celda(fr2, 4, d['tiempo_entrega'], merge_hasta=6)
        fr2 += 1
    if d.get('condicion_pago'):
        ws.merge_cells(f'B{fr2}:C{fr2}')
        celda(fr2, 2, "Condición de Pago:", bold=True, color_fondo=gris_claro)
        celda(fr2, 4, d['condicion_pago'], merge_hasta=6)
        fr2 += 1

    fr3 = fr2 + 3
    celda(fr3, 4, "OSCAR ELERA", bold=True, alin="center", merge_hasta=6)
    celda(fr3+1, 4, "DOBLE E PUBLICIDAD Y DISEÑO", alin="center", merge_hasta=6)
    celda(fr3+2, 4, "Oficina 2661221 / 988455940", alin="center", merge_hasta=6)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def generar_pdf(d):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    azul = colors.HexColor('#1B2A4A')
    azul_m = colors.HexColor('#2E4F8C')
    amarillo = colors.HexColor('#FFC107')
    gris = colors.HexColor('#F5F5F5')
    story = []

    header_data = [
        [Paragraph('<font color="white"><b>DOBLE E PUBLICIDAD Y DISEÑO</b></font>',
                   ParagraphStyle('h', fontName='Helvetica-Bold', fontSize=14, alignment=TA_CENTER)),
         Paragraph('<font color="white"><b>N° Cotización</b></font>',
                   ParagraphStyle('h', fontName='Helvetica-Bold', fontSize=10, alignment=TA_CENTER)),
         Paragraph(f'<b>{d["numero"]}</b>', ParagraphStyle('h', fontName='Helvetica-Bold', fontSize=14, alignment=TA_CENTER))],
        ['',
         Paragraph('<font color="white"><b>Fecha</b></font>',
                   ParagraphStyle('h', fontName='Helvetica-Bold', fontSize=10, alignment=TA_CENTER)),
         Paragraph(d['fecha'], ParagraphStyle('h', fontSize=10, alignment=TA_CENTER))]
    ]
    t = Table(header_data, colWidths=[10*cm, 3.5*cm, 4*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,1), azul),
        ('BACKGROUND', (1,0), (1,0), azul_m),
        ('BACKGROUND', (2,0), (2,0), amarillo),
        ('BACKGROUND', (1,1), (1,1), azul_m),
        ('BACKGROUND', (2,1), (2,1), gris),
        ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.3*cm))

    emisor_data = [
        [Paragraph('<b>RUC:</b>', ParagraphStyle('n', fontSize=9)), '10011273659'],
        [Paragraph('<b>Dirección:</b>', ParagraphStyle('n', fontSize=9)), 'Los Diamantes 387 int. 303 La Victoria Lima'],
        [Paragraph('<b>Teléfono:</b>', ParagraphStyle('n', fontSize=9)), 'Oficina 2661221 / 988455940'],
    ]
    te = Table(emisor_data, colWidths=[3*cm, 14.5*cm])
    te.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), gris),
        ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(te)
    story.append(Spacer(1, 0.3*cm))

    cliente_data = [
        [Paragraph('<font color="white"><b>DATOS DEL CLIENTE</b></font>',
                   ParagraphStyle('h', fontName='Helvetica-Bold', fontSize=10, alignment=TA_CENTER)), ''],
        [Paragraph('<b>Cliente:</b>', ParagraphStyle('n', fontSize=9)), d.get('cliente','')],
        [Paragraph('<b>RUC:</b>', ParagraphStyle('n', fontSize=9)), d.get('ruc','')],
        [Paragraph('<b>Dirección:</b>', ParagraphStyle('n', fontSize=9)), d.get('direccion','')],
        [Paragraph('<b>R. Social:</b>', ParagraphStyle('n', fontSize=9)), d.get('razon_social','')],
        [Paragraph('<b>Contacto:</b>', ParagraphStyle('n', fontSize=9)), d.get('contacto','')],
        [Paragraph('<b>Teléfonos:</b>', ParagraphStyle('n', fontSize=9)), d.get('telefonos','')],
    ]
    tc = Table(cliente_data, colWidths=[3*cm, 14.5*cm])
    tc.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), azul),
        ('SPAN', (0,0), (-1,0)),
        ('BACKGROUND', (0,1), (0,-1), gris),
        ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(tc)
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph("En atención a su gentil solicitud, es grato cotizar a ustedes el siguiente trabajo:",
                            ParagraphStyle('n', fontSize=9, backColor=colors.HexColor('#FFF8E1'))))
    story.append(Spacer(1, 0.2*cm))

    items_header = [
        Paragraph('<font color="white"><b>ITEM</b></font>', ParagraphStyle('c', fontSize=9, alignment=TA_CENTER)),
        Paragraph('<font color="white"><b>CANT.</b></font>', ParagraphStyle('c', fontSize=9, alignment=TA_CENTER)),
        Paragraph('<font color="white"><b>DETALLE</b></font>', ParagraphStyle('c', fontSize=9, alignment=TA_CENTER)),
        Paragraph('<font color="white"><b>PRECIO UNIT.</b></font>', ParagraphStyle('c', fontSize=9, alignment=TA_CENTER)),
        Paragraph('<font color="white"><b>V. TOTAL</b></font>', ParagraphStyle('c', fontSize=9, alignment=TA_CENTER)),
    ]
    items_data = [items_header]
    for i, item in enumerate(d['items'], 1):
        total_item = item.get('cantidad', 0) * item.get('precio', 0)
        items_data.append([str(i), str(item.get('cantidad','')), item.get('detalle',''),
                           f"S/ {item.get('precio',0):,.2f}", f"S/ {total_item:,.2f}"])

    ti = Table(items_data, colWidths=[1*cm, 1.5*cm, 9.5*cm, 2.5*cm, 3*cm])
    row_colors = [('BACKGROUND', (0,0), (-1,0), azul_m)]
    for i in range(1, len(items_data)):
        bg = colors.white if i % 2 == 1 else gris
        row_colors.append(('BACKGROUND', (0,i), (-1,i), bg))
    ti.setStyle(TableStyle(row_colors + [
        ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (0,0), (1,-1), 'CENTER'),
        ('ALIGN', (3,0), (-1,-1), 'RIGHT'),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(ti)
    story.append(Spacer(1, 0.2*cm))

    totales = [
        ['', Paragraph('<b>V. VENTA (SIN IGV)</b>', ParagraphStyle('r', fontSize=10, alignment=TA_RIGHT)),
         Paragraph(f'<b>S/ {d["subtotal"]:,.2f}</b>', ParagraphStyle('r', fontSize=10, alignment=TA_RIGHT))],
        ['', Paragraph('<b>IGV (18%)</b>', ParagraphStyle('r', fontSize=10, alignment=TA_RIGHT)),
         Paragraph(f'S/ {d["igv"]:,.2f}', ParagraphStyle('r', fontSize=10, alignment=TA_RIGHT))],
        ['', Paragraph('<font color="white"><b>PRECIO TOTAL S/.</b></font>',
                       ParagraphStyle('r', fontName='Helvetica-Bold', fontSize=11, alignment=TA_RIGHT)),
         Paragraph(f'<b>S/ {d["total"]:,.2f}</b>',
                   ParagraphStyle('r', fontName='Helvetica-Bold', fontSize=11, alignment=TA_RIGHT))],
    ]
    tt = Table(totales, colWidths=[7.5*cm, 6*cm, 4*cm])
    tt.setStyle(TableStyle([
        ('BACKGROUND', (1,0), (-1,0), gris),
        ('BACKGROUND', (1,1), (-1,1), gris),
        ('BACKGROUND', (1,2), (1,2), azul),
        ('BACKGROUND', (2,2), (2,2), amarillo),
        ('BOX', (1,0), (-1,-1), 0.5, colors.grey),
        ('INNERGRID', (1,0), (-1,-1), 0.5, colors.grey),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 8),
    ]))
    story.append(tt)

    if d.get('tiempo_entrega') or d.get('condicion_pago'):
        story.append(Spacer(1, 0.4*cm))
        cond_data = []
        if d.get('tiempo_entrega'):
            cond_data.append([Paragraph('<b>Tiempo de Entrega:</b>', ParagraphStyle('n', fontSize=9)), d['tiempo_entrega']])
        if d.get('condicion_pago'):
            cond_data.append([Paragraph('<b>Condición de Pago:</b>', ParagraphStyle('n', fontSize=9)), d['condicion_pago']])
        tc2 = Table(cond_data, colWidths=[4*cm, 13.5*cm])
        tc2.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,-1), gris),
            ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]))
        story.append(tc2)

    story.append(Spacer(1, 1*cm))
    firma_data = [
        [Paragraph('<b>OSCAR ELERA</b>', ParagraphStyle('c', fontSize=10, alignment=TA_CENTER))],
        [Paragraph('DOBLE E PUBLICIDAD Y DISEÑO', ParagraphStyle('c', fontSize=9, alignment=TA_CENTER))],
        [Paragraph('Oficina 2661221 / 988455940', ParagraphStyle('c', fontSize=9, alignment=TA_CENTER))],
    ]
    tf = Table(firma_data, colWidths=[17.5*cm])
    story.append(tf)
    doc.build(story)
    buf.seek(0)
    return buf

if __name__ == '__main__':
    init_db()
    print("\n✅ Cotizador Doble E listo en http://localhost:5000\n")
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
